import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from alphabet import num_to_akuk
import tkinter as tk
from tkinter import messagebox
from tkinter import *

# Function to handle the save button click
def save_to_excel():
    # Get the input values
    num = float(num_entry.get())
    meaning = meaning_entry.get()

    # Load the workbook
    file_name = 'akuk_words.xlsx'
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Get the active worksheet
    ws = wb.active

    # Check if the number already exists in the worksheet
    num_col = ws['A']
    num_values = [cell.value for cell in num_col]
    if num in num_values:
        # Ask for confirmation before updating
        confirm = messagebox.askyesno("Update Existing Entry", f"Akuk word for {num} already exists. Do you want to update it?")
        if not confirm:
            return

        # Update the existing row
        row_num = num_values.index(num) + 1
        akuk_col = ws[f"B{row_num}"]
        meaning_col = ws[f"C{row_num}"]
    else:
        # Add a new row
        row_num = ws.max_row + 1
        num_col = ws.cell(row=row_num, column=1, value=num)
        akuk_col = ws.cell(row=row_num, column=2, value=num_to_akuk(num))
        meaning_col = ws.cell(row=row_num, column=3, value=meaning)

    # Apply font style to the header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Save the workbook
    wb.save(file_name)
    messagebox.showinfo("Save Success", "The entry has been saved to the file.")

# Create the GUI
root = tk.Tk()
root.title("Add Vocabulary")
root.geometry("800x600")

#image
p1 = PhotoImage(file = 'icon.png')
root.iconphoto(False, p1)

# Add the input fields
num_label = tk.Label(root, text="Numeric Value:")
num_label.pack()
num_entry = tk.Entry(root)
num_entry.pack()

meaning_label = tk.Label(root, text="Meaning:")
meaning_label.pack()
meaning_entry = tk.Entry(root)
meaning_entry.pack()

# Add the save button
save_button = tk.Button(root, text="Save", command=save_to_excel)
save_button.pack()

# Start the main loop
root.mainloop()
